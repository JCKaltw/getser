import openpyxl
import csv
import re
import sys
import os

def is_hex(s):
    return bool(re.match("^[0-9A-Fa-f]{8,}$", s))

def col_letter(col_num):
    column_letter = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        column_letter = chr(65 + remainder) + column_letter
    return column_letter

def process_worksheet(ws):
    data = []
    found = False
    hex_row_num = None
    for row_num, row in enumerate(ws.iter_rows(), start=1):
        for col_num, cell in enumerate(row, start=1):
            if is_hex(str(cell.value)):
                col1 = col_letter(col_num)
                col2 = col_letter(col_num + 1)
                col3 = col_letter(col_num + 2)
                data.append([cell.value, ws[f"{col2}{row_num}"].value, ws[f"{col3}{row_num}"].value])
                found = True
                hex_row_num = row_num
                break
        if found:
            break

    if hex_row_num is not None:
        for row in ws.iter_rows(min_row=hex_row_num + 1):
            values = [cell.value for col_num, cell in enumerate(row, start=1) if col_letter(col_num) in {col1, col2, col3}]
            if not any(values):
                break
            data.append(values)

    return data

def main(input_file):
    wb = openpyxl.load_workbook(input_file)
    log_file = open("getser.log", "w")

    script_dir = os.path.dirname(os.path.realpath(__file__))
    data_dir = os.path.join(script_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    log_file.write(f"Input file: {input_file}\n")
    log_file.write(f"Script directory: {script_dir}\n")
    log_file.write(f"Data directory: {data_dir}\n")

    for ws in wb.worksheets:
        log_file.write(f"Processing worksheet: {ws.title}\n")
        data = process_worksheet(ws)
        if not data:
            log_file.write(f"Error: Worksheet {ws.title} - Could not locate 3 columns with hex data.\n")
            print(f"{ws.title}: skipped ERROR")
            continue

        csv_file_path = os.path.join(data_dir, f"{ws.title}.csv")
        with open(csv_file_path, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["generator-serial", "sensor-serial", "pairing-serial"])
            writer.writerows(data)
        
        log_file.write(f"Created CSV file: {csv_file_path}\n")
        print(f"{ws.title}: extracted OK")

    log_file.close()

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python script_name.py input_file.xlsx")
        sys.exit(1)

    main(sys.argv[1])
